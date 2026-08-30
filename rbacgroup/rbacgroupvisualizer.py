#!/usr/bin/env python3
from __future__ import annotations
import argparse
import math
import re
import shutil
from pathlib import Path
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill


REQUIRED_COLUMNS = {
    "User Clusters": {"ClusterId", "User", "UserId"},
    "Role Candidates": {
        "ClusterId", "CandidateGroup", "Prevalence", "GroupId"
    },
    "Members": {"MemberId", "GroupId"},
}

OVERVIEW_FILENAME = "rbac_overview.png"
PNG_DPI = 180
EXCEL_IMAGE_MAX_WIDTH = 1500


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate RBAC heatmaps and an overview from rbacgroup.py output."
    )
    parser.add_argument("workbook", type=Path, help="Input Excel workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="PNG destination (default: <input-stem>_visuals)",
    )
    parser.add_argument(
        "--workbook-output",
        type=Path,
        help="Visualized workbook destination (default: <input-stem>_visualized.xlsx)",
    )
    parser.add_argument(
        "--no-workbook",
        action="store_true",
        help="Create PNG files only; do not create a visualized workbook copy.",
    )
    parser.add_argument(
        "--dpi", type=int, default=PNG_DPI, help=f"PNG resolution (default: {PNG_DPI})"
    )
    return parser.parse_args()


def safe_filename(value: object) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value)).strip("._")
    return text or "unknown"


def sheet_frame(path: Path, sheet: str) -> pd.DataFrame:
    frame = pd.read_excel(path, sheet_name=sheet, dtype=object)
    missing = REQUIRED_COLUMNS[sheet] - set(frame.columns)
    if missing:
        raise ValueError(
            f"Sheet '{sheet}' is missing required columns: {', '.join(sorted(missing))}"
        )
    return frame


def load_analysis(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    with pd.ExcelFile(path) as excel:
        missing_sheets = set(REQUIRED_COLUMNS) - set(excel.sheet_names)
    if missing_sheets:
        raise ValueError(
            "Workbook is not a compatible sailfail2.py report. Missing sheets: "
            + ", ".join(sorted(missing_sheets))
        )

    clusters = sheet_frame(path, "User Clusters").dropna(subset=["ClusterId", "UserId"])
    candidates = sheet_frame(path, "Role Candidates").dropna(
        subset=["ClusterId", "GroupId"]
    )
    members = sheet_frame(path, "Members").dropna(subset=["MemberId", "GroupId"])

    for frame in (clusters, candidates):
        frame["ClusterId"] = pd.to_numeric(frame["ClusterId"], errors="raise").astype(int)
    candidates["Prevalence"] = pd.to_numeric(
        candidates["Prevalence"], errors="coerce"
    ).fillna(0.0)

    # IDs must be compared as strings; Excel sometimes infers numeric-looking IDs.
    clusters["UserId"] = clusters["UserId"].astype(str)
    candidates["GroupId"] = candidates["GroupId"].astype(str)
    members["MemberId"] = members["MemberId"].astype(str)
    members["GroupId"] = members["GroupId"].astype(str)

    return clusters, candidates, members


def cluster_matrix(
    cluster_id: int,
    clusters: pd.DataFrame,
    candidates: pd.DataFrame,
    membership_pairs: set[tuple[str, str]],
) -> pd.DataFrame:
    users = (
        clusters.loc[clusters["ClusterId"] == cluster_id, ["UserId", "User"]]
        .drop_duplicates("UserId")
        .sort_values(["User", "UserId"], kind="stable")
    )
    groups = (
        candidates.loc[
            candidates["ClusterId"] == cluster_id,
            ["GroupId", "CandidateGroup", "Prevalence"],
        ]
        .drop_duplicates("GroupId")
        .sort_values(["Prevalence", "CandidateGroup"], ascending=[False, True])
    )

    # CandidateGroup already disambiguates duplicate display names when needed.
    row_labels = [str(v) for v in users["User"]]
    col_labels = [str(v) for v in groups["CandidateGroup"]]
    values = [
        [1 if (str(uid), str(gid)) in membership_pairs else 0 for gid in groups["GroupId"]]
        for uid in users["UserId"]
    ]
    return pd.DataFrame(values, index=row_labels, columns=col_labels, dtype=int)


def tick_step(count: int, maximum_labels: int) -> int:
    return max(1, math.ceil(count / maximum_labels))


def save_cluster_heatmap(matrix: pd.DataFrame, cluster_id: int, output: Path,
                         dpi: int) -> None:
    if matrix.empty or not len(matrix.columns):
        return

    width = min(30, max(8, 2.5 + 0.48 * len(matrix.columns)))
    height = min(40, max(5, 2.5 + 0.28 * len(matrix.index)))
    fig, ax = plt.subplots(figsize=(width, height))

    sns.heatmap(
        matrix,
        cmap=sns.color_palette(["#F3F4F6", "#2F75B5"]),
        vmin=0,
        vmax=1,
        linewidths=0.35,
        linecolor="white",
        cbar=False,
        ax=ax,
    )
    ax.set_title(
        f"RBAC candidate cluster {cluster_id} — "
        f"{len(matrix.index)} users × {len(matrix.columns)} candidate groups",
        fontsize=14,
        fontweight="bold",
        pad=16,
    )
    ax.set_xlabel("Candidate role groups")
    ax.set_ylabel("Users")

    x_step = tick_step(len(matrix.columns), 45)
    y_step = tick_step(len(matrix.index), 70)
    ax.set_xticks([i + 0.5 for i in range(0, len(matrix.columns), x_step)])
    ax.set_xticklabels(matrix.columns[::x_step], rotation=55, ha="right", fontsize=8)
    ax.set_yticks([i + 0.5 for i in range(0, len(matrix.index), y_step)])
    ax.set_yticklabels(matrix.index[::y_step], rotation=0, fontsize=8)

    missing = int(matrix.size - matrix.to_numpy().sum())
    ax.text(
        1.0,
        1.01,
        f"Blue = member  |  Gray = missing  |  Missing assignments: {missing}",
        transform=ax.transAxes,
        ha="right",
        va="bottom",
        fontsize=9,
        color="#555555",
    )
    fig.tight_layout()
    fig.savefig(output, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def overview_metrics(
    clusters: pd.DataFrame,
    candidates: pd.DataFrame,
    matrices: dict[int, pd.DataFrame],
) -> pd.DataFrame:
    rows = []
    for cluster_id, matrix in matrices.items():
        candidate_rows = candidates[candidates["ClusterId"] == cluster_id]
        missing = int(matrix.size - matrix.to_numpy().sum()) if matrix.size else 0
        rows.append(
            {
                "ClusterId": cluster_id,
                "Users": int(clusters.loc[
                    clusters["ClusterId"] == cluster_id, "UserId"
                ].nunique()),
                "CandidateGroups": int(candidate_rows["GroupId"].nunique()),
                "MeanPrevalence": float(candidate_rows["Prevalence"].mean()),
                "MissingAssignments": missing,
            }
        )
    return pd.DataFrame(rows).sort_values("Users", ascending=True)


def save_overview(metrics: pd.DataFrame, output: Path, dpi: int) -> None:
    if metrics.empty:
        raise ValueError("No role candidates were found; an RBAC overview cannot be created.")

    count = len(metrics)
    height = max(5.5, min(30, 2.5 + 0.48 * count))
    fig, axes = plt.subplots(
        1, 3, figsize=(16, height), sharey=True,
        gridspec_kw={"width_ratios": [1.25, 1.25, 1.5]},
    )
    labels = [f"Cluster {value}" for value in metrics["ClusterId"]]
    y = list(range(count))

    panels = [
        (axes[0], "Users", "Users", "#2F75B5"),
        (axes[1], "CandidateGroups", "Candidate groups", "#70AD47"),
        (axes[2], "MissingAssignments", "Missing memberships", "#C55A11"),
    ]
    for ax, column, title, color in panels:
        values = metrics[column].tolist()
        bars = ax.barh(y, values, color=color, alpha=0.9)
        ax.set_title(title, fontweight="bold")
        ax.set_xlabel("Count")
        ax.grid(axis="x", alpha=0.2)
        ax.spines[["top", "right", "left"]].set_visible(False)
        ax.bar_label(bars, padding=3, fontsize=9)

    axes[0].set_yticks(y, labels)
    for idx, prevalence in enumerate(metrics["MeanPrevalence"]):
        axes[2].text(
            1.0,
            idx,
            f"  mean prevalence {prevalence:.0%}",
            transform=axes[2].get_yaxis_transform(),
            va="center",
            ha="left",
            fontsize=8,
            color="#555555",
        )

    fig.suptitle(
        "RBAC discovery overview",
        fontsize=18,
        fontweight="bold",
        x=0.04,
        y=0.985,
        ha="left",
    )
    fig.text(
        0.04,
        0.895,
        "Candidate populations derived from Jaccard similarity; role groups derived from cluster prevalence",
        fontsize=10,
        color="#555555",
        ha="left",
    )
    fig.tight_layout(rect=(0.03, 0.02, 0.97, 0.84))
    fig.savefig(output, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def scale_excel_image(image: ExcelImage, max_width: int = EXCEL_IMAGE_MAX_WIDTH) -> None:
    if image.width > max_width:
        ratio = max_width / image.width
        image.width = int(image.width * ratio)
        image.height = int(image.height * ratio)


def add_title(ws, title: str, subtitle: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="555555")
    ws.column_dimensions["A"].width = 28


def unique_sheet_name(workbook, desired: str) -> str:
    base = desired[:31]
    if base not in workbook.sheetnames:
        return base
    suffix = 2
    while True:
        candidate = f"{base[:27]} {suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


def build_visualized_workbook(
    source: Path,
    destination: Path,
    overview_path: Path,
    cluster_paths: dict[int, Path],
) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() == destination.resolve():
        raise ValueError("Workbook output must differ from the source workbook.")
    shutil.copy2(source, destination)
    workbook = load_workbook(destination)

    overview_sheet = workbook.create_sheet(unique_sheet_name(workbook, "RBAC Visuals"), 1)
    add_title(
        overview_sheet,
        "RBAC discovery overview",
        "Use the detailed data sheets to validate every candidate role and exception.",
    )
    overview_image = ExcelImage(overview_path)
    scale_excel_image(overview_image)
    overview_sheet.add_image(overview_image, "A4")

    for cluster_id, image_path in sorted(cluster_paths.items()):
        sheet = workbook.create_sheet(
            unique_sheet_name(workbook, f"Heatmap C{cluster_id:03d}")
        )
        add_title(
            sheet,
            f"Candidate cluster {cluster_id}",
            "Blue cells are current memberships; gray cells are reviewable gaps.",
        )
        image = ExcelImage(image_path)
        scale_excel_image(image)
        sheet.add_image(image, "A4")

    workbook.save(destination)


def main() -> None:
    args = parse_args()
    source = args.workbook.expanduser().resolve()
    output_dir = (
        args.output_dir.expanduser().resolve()
        if args.output_dir
        else source.with_name(f"{source.stem}_visuals")
    )
    workbook_output = (
        args.workbook_output.expanduser().resolve()
        if args.workbook_output
        else source.with_name(f"{source.stem}_visualized.xlsx")
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    clusters, candidates, members = load_analysis(source)
    if clusters.empty:
        raise ValueError("The User Clusters sheet contains no candidate clusters.")
    if candidates.empty:
        raise ValueError("The Role Candidates sheet contains no candidate groups.")

    membership_pairs = set(zip(members["MemberId"], members["GroupId"]))
    cluster_ids = sorted(set(clusters["ClusterId"]) & set(candidates["ClusterId"]))
    if not cluster_ids:
        raise ValueError("No ClusterId is shared by User Clusters and Role Candidates.")

    matrices = {
        cluster_id: cluster_matrix(
            cluster_id, clusters, candidates, membership_pairs
        )
        for cluster_id in cluster_ids
    }
    cluster_paths = {}
    for cluster_id, matrix in matrices.items():
        path = output_dir / f"cluster_{safe_filename(cluster_id)}_heatmap.png"
        save_cluster_heatmap(matrix, cluster_id, path, args.dpi)
        cluster_paths[cluster_id] = path

    overview_path = output_dir / OVERVIEW_FILENAME
    metrics = overview_metrics(clusters, candidates, matrices)
    save_overview(metrics, overview_path, args.dpi)

    if not args.no_workbook:
        build_visualized_workbook(
            source, workbook_output, overview_path, cluster_paths
        )

    print(f"Overview: {overview_path}")
    print(f"Cluster heatmaps: {len(cluster_paths)} in {output_dir}")
    if not args.no_workbook:
        print(f"Visualized workbook: {workbook_output}")


if __name__ == "__main__":
    main()
