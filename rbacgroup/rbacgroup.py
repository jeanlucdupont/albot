#!/usr/bin/env python3
import os
import itertools
import time
import warnings
from collections import defaultdict, deque
from datetime import datetime, timezone
from dotenv import load_dotenv
import msal
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from mlxtend.frequent_patterns import association_rules, fpgrowth
from mlxtend.preprocessing import TransactionEncoder


load_dotenv()
TENANT_ID     = os.environ.get("TENANT_ID")
CLIENT_ID     = os.environ.get("CLIENT_ID")
CLIENT_SECRET = os.environ.get("CLIENT_SECRET")
SIMILARITY_THRESHOLD = 0.85          # group-to-group comparison
SIMILARITY_METRIC = "overlap"
USER_JACCARD_THRESHOLD = 0.75
MIN_SHARED_GROUPS = 3
MIN_ROLE_USERS = 5
ROLE_GROUP_PREVALENCE = 0.80
MIN_ITEMSET_SUPPORT = 0.005          # also bounded by MIN_ROLE_USERS below
MIN_RULE_CONFIDENCE = 0.80
MIN_RULE_LIFT = 1.50
MAX_ITEMSET_LENGTH = 5
RBAC_SECURITY_GROUPS_ONLY = True
LIFECYCLE_REVIEW_DAYS = 365          # old lifecycle metadata; NOT an activity signal
TRANSITIVE = False                   # True -> expand nested-group membership in Members
DROP_UBIQUITOUS_ABOVE = 0.95         # drop groups present in >95% of users pre-mining
MAX_OUTPUT_ROWS = 50_000             # hard cap for Frequent Bundles / Association Rules
OUTFILE = "rbacgroup.xlsx"
GRAPH = "https://graph.microsoft.com/v1.0"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPE = ["https://graph.microsoft.com/.default"]
NOW = datetime.now(timezone.utc)
MAX_GRAPH_RETRIES = 6

def get_token():
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
        raise RuntimeError("Missing TENANT_ID / CLIENT_ID / CLIENT_SECRET in env/.env")
    app = msal.ConfidentialClientApplication(
        CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET
    )
    result = app.acquire_token_for_client(scopes=SCOPE)
    if "access_token" not in result:
        raise RuntimeError(f"Auth failed: {result.get('error_description', result)}")
    return result["access_token"]
def graph_get_all(url, headers):
    """Handle Graph paging and transient throttling. Raises on final error."""
    items = []
    while url:
        for attempt in range(MAX_GRAPH_RETRIES + 1):
            r = requests.get(url, headers=headers, timeout=60)
            if r.status_code not in (429, 503, 504):
                break
            if attempt == MAX_GRAPH_RETRIES:
                break
            retry_after = r.headers.get("Retry-After")
            try:
                delay = float(retry_after) if retry_after else min(2 ** attempt, 60)
            except ValueError:
                delay = min(2 ** attempt, 60)
            print(f"  Graph returned HTTP {r.status_code}; retrying in {delay:g}s")
            time.sleep(delay)
        r.raise_for_status()
        data = r.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items
def graph_get_all_safe(url, headers):
    """Like graph_get_all but returns (items, error). Never raises on 403/permission."""
    try:
        return graph_get_all(url, headers), None
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else "?"
        return [], f"HTTP {code}"

GROUP_SELECT = (
    "id,displayName,groupTypes,securityEnabled,mailEnabled,"
    "onPremisesSyncEnabled,membershipRule,isAssignableToRole,"
    "createdDateTime,renewedDateTime,visibility,resourceProvisioningOptions"
)
def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None
def days_since(dt):
    return (NOW - dt).days if dt else None
def classify_group(g):
    if "Unified" in (g.get("groupTypes") or []):
        return "Microsoft 365"
    if g.get("securityEnabled") and not g.get("mailEnabled"):
        return "Security"
    if g.get("mailEnabled") and not g.get("securityEnabled"):
        return "Distribution"
    return "Mail-enabled security"
def short_type(odata_type):
    """'#microsoft.graph.user' -> 'user'."""
    return (odata_type or "").rsplit(".", 1)[-1] or "unknown"
def collect_groups(headers):
    groups = graph_get_all(
        f"{GRAPH}/groups?$select={GROUP_SELECT}&$top=999",
        headers,
    )
    member_ep = "transitiveMembers" if TRANSITIVE else "members"
    print("Listing groups")
    data = {}
    for g in groups:
        gid = g["id"]
        # Retrieve users, groups, devices, and other standard members
        members = graph_get_all(
            f"{GRAPH}/groups/{gid}/{member_ep}"
            f"?$select=id,displayName,userPrincipalName&$top=999",
            headers,
        )
        # Retrieve service principals separately because the generic endpoint
        # may not return them.
        service_principals = graph_get_all(
            f"{GRAPH}/groups/{gid}/{member_ep}/microsoft.graph.servicePrincipal"
            f"?$select=id,displayName,appDisplayName,servicePrincipalType&$top=999",
            headers,
        )
        # Ensure service principals have a type, since the cast endpoint may
        # not return @odata.type.
        for sp in service_principals:
            sp.setdefault("@odata.type", "#microsoft.graph.servicePrincipal")
        # Merge both results and remove duplicates by object ID
        members_by_id = {
            m["id"]: m
            for m in members + service_principals
            if m.get("id")
        }
        members = list(members_by_id.values())
        member_ids = set()
        member_objs = {}
        child_groups = set()
        for m in members:
            mid = m.get("id")
            if not mid:
                continue
            mtype = short_type(m.get("@odata.type"))
            member_ids.add(mid)
            member_objs[mid] = {
                "name": (
                    m.get("userPrincipalName")
                    or m.get("displayName")
                    or m.get("appDisplayName")
                    or mid
                ),
                "type": mtype,
            }
            if mtype == "group":
                child_groups.add(mid)
        owners = graph_get_all(
            f"{GRAPH}/groups/{gid}/owners"
            f"?$select=id,displayName,userPrincipalName&$top=999",
            headers,
        )
        owner_names = [
            o.get("userPrincipalName")
            or o.get("displayName")
            or o.get("id")
            for o in owners
        ]
        created = parse_dt(g.get("createdDateTime"))
        renewed = parse_dt(g.get("renewedDateTime"))
        last_touch = max(
            [d for d in (created, renewed) if d],
            default=None,
        )
        data[gid] = {
            "name": g.get("displayName", ""),
            "type": classify_group(g),
            "security_enabled": bool(g.get("securityEnabled")),
            "members": member_ids,
            "member_objs": member_objs,
            "child_groups": child_groups,
            "owners": owner_names,
            "synced": bool(g.get("onPremisesSyncEnabled")),
            "dynamic": bool(g.get("membershipRule")),
            "role_assignable": bool(g.get("isAssignableToRole")),
            "teams_connected": (
                "Team" in (g.get("resourceProvisioningOptions") or [])
            ),
            "visibility": g.get("visibility") or "",
            "created": created,
            "renewed": renewed,
            # This is lifecycle metadata, not evidence of access/activity.
            "lifecycle_age_days": days_since(last_touch),
            # Filled later
            "parents": set(),
            "direct_roles": [],
        }
        print(
            f"  {g.get('displayName', ''):46.46}  "
            f"members={len(member_ids):<4} "
            f"owners={len(owner_names)} "
            f"teams={'Y' if data[gid]['teams_connected'] else '-'}"
        )
    return data
def build_nesting(data):
    """Populate each group's 'parents' from child relationships."""
    for gid, g in data.items():
        for child in g["child_groups"]:
            if child in data:
                data[child]["parents"].add(gid)
def ancestors(gid, data):
    """All groups this group is transitively a member of (upward walk)."""
    seen, q = set(), deque(data[gid]["parents"])
    while q:
        p = q.popleft()
        if p in seen or p not in data:
            continue
        seen.add(p)
        q.extend(data[p]["parents"])
    return seen

def collect_roles(headers, data):
    """
    Fill data[gid]['direct_roles'] with (roleName, kind).
      kind = 'Active'        -> standing role assignment
      kind = 'Eligible (PIM)'-> eligible via PIM
    Returns (roles_available: bool, note: str).
    """
    active, err1 = graph_get_all_safe(
        f"{GRAPH}/roleManagement/directory/roleAssignments?$expand=roleDefinition&$top=100",
        headers,
    )
    eligible, err2 = graph_get_all_safe(
        f"{GRAPH}/roleManagement/directory/roleEligibilityScheduleInstances"
        f"?$expand=roleDefinition&$top=100",
        headers,
    )
    if err1 or err2:
        return False, f"Role lookup skipped ({err1 or err2}). " \
                      f"Add RoleManagement.Read.Directory and admin-consent."
    def role_name(rec):
        rd = rec.get("roleDefinition") or {}
        return rd.get("displayName") or rec.get("roleDefinitionId", "?")
    for rec in active:
        pid = rec.get("principalId")
        if pid in data:
            data[pid]["direct_roles"].append((role_name(rec), "Active"))
    for rec in eligible:
        pid = rec.get("principalId")
        if pid in data:
            data[pid]["direct_roles"].append((role_name(rec), "Eligible (PIM)"))
    return True, "OK"
def effective_roles(gid, data):
    """
    Return direct directory-role assignments only.
    Nested group membership must not be treated as effective Entra directory-
    role inheritance. Role-assignable groups restrict active group nesting,
    and PIM-eligible group relationships require separate effective-access
    evaluation.
    """
    return [(rn, kind, "Direct") for rn, kind in data[gid]["direct_roles"]]
def role_summary(gid, data):
    roles = effective_roles(gid, data)
    if not roles:
        return ""
    parts = []
    for rn, kind, src in roles:
        tag = rn if src == "Direct" else f"{rn} ({src})"
        parts.append(f"{kind}: {tag}")
    # de-dup while preserving order
    seen, uniq = set(), []
    for p in parts:
        if p not in seen:
            seen.add(p); uniq.append(p)
    return " | ".join(uniq)
def is_privileged(gid, data):
    return bool(effective_roles(gid, data))

def similarity(a, b, metric):
    if not a or not b:
        return 0.0
    inter = len(a & b)
    return inter / len(a | b) if metric == "jaccard" else inter / min(len(a), len(b))
def find_identical(data):
    buckets = defaultdict(list)
    for gid, g in data.items():
        if g["members"]:
            buckets[frozenset(g["members"])].append(gid)
    return [gids for gids in buckets.values() if len(gids) > 1]
def find_similar_and_subsets(data, threshold, metric):
    items = [(gid, g) for gid, g in data.items() if g["members"]]
    similar, subsets = [], []
    for (id_a, a), (id_b, b) in itertools.combinations(items, 2):
        ma, mb = a["members"], b["members"]
        shared = len(ma & mb)
        if not shared:
            continue
        s = similarity(ma, mb, metric)
        if s >= threshold:
            similar.append((id_a, id_b, s, shared))
        if ma < mb:
            subsets.append((id_a, id_b, shared))       # A subset of B
        elif mb < ma:
            subsets.append((id_b, id_a, shared))
    return (sorted(similar, key=lambda x: x[2], reverse=True), subsets)
def needs_lifecycle_review(g):
    age = g["lifecycle_age_days"]
    return age is not None and age >= LIFECYCLE_REVIEW_DAYS
def cleanup_reasons(g):
    r = []
    if not g["owners"]:
        r.append("No owner")
    if not g["members"]:
        r.append("Empty")
    if needs_lifecycle_review(g):
        r.append(f"Lifecycle review (>{LIFECYCLE_REVIEW_DAYS}d; not activity-based)")
    return r
def rbac_group_ids(data):
    """Groups eligible to influence RBAC pattern discovery."""
    return {
        gid for gid, g in data.items()
        if not RBAC_SECURITY_GROUPS_ONLY or g["security_enabled"]
    }
def build_user_group_sets(data):
    """Return direct/transitive (per config) user memberships using group IDs."""
    allowed = rbac_group_ids(data)
    memberships = defaultdict(set)
    names = {}
    for gid, g in data.items():
        if gid not in allowed:
            continue
        for mid, member in g["member_objs"].items():
            if member.get("type") != "user":
                continue
            memberships[mid].add(gid)
            names[mid] = member.get("name") or mid
    return dict(memberships), names
def jaccard(a, b):
    union = a | b
    return len(a & b) / len(union) if union else 1.0
def find_similar_users(user_groups):
    """Find meaningful pairs without comparing every user to every other."""
    group_users = defaultdict(list)
    for uid, groups in user_groups.items():
        for gid in groups:
            group_users[gid].append(uid)
    shared_counts = defaultdict(int)
    for users in group_users.values():
        for u1, u2 in itertools.combinations(sorted(users), 2):
            shared_counts[(u1, u2)] += 1
    pairs = []
    for (u1, u2), shared_count in shared_counts.items():
        if shared_count < MIN_SHARED_GROUPS:
            continue
        g1 = user_groups[u1]
        g2 = user_groups[u2]
        inter = g1 & g2
        score = jaccard(g1, g2)
        if score >= USER_JACCARD_THRESHOLD:
            pairs.append((u1, u2, score, inter, g1 - g2, g2 - g1))
    return sorted(pairs, key=lambda x: x[2], reverse=True)
def cluster_similar_users(user_groups, similar_users):
    """
    Create complete-link-style clusters: every admitted user must meet the
    Jaccard threshold and minimum intersection against every cluster member.
    """
    eligible = set()
    for u1, u2, *_ in similar_users:
        eligible.update((u1, u2))
    ordered = sorted(eligible, key=lambda u: (-len(user_groups[u]), u))
    clusters = []
    for user in ordered:
        best_idx, best_score = None, -1.0
        for idx, cluster in enumerate(clusters):
            scores = []
            valid = True
            for member in cluster:
                shared = user_groups[user] & user_groups[member]
                score = jaccard(user_groups[user], user_groups[member])
                if len(shared) < MIN_SHARED_GROUPS or score < USER_JACCARD_THRESHOLD:
                    valid = False
                    break
                scores.append(score)
            if valid and scores and min(scores) > best_score:
                best_idx, best_score = idx, min(scores)
        if best_idx is None:
            clusters.append([user])
        else:
            clusters[best_idx].append(user)
    return [c for c in clusters if len(c) >= MIN_ROLE_USERS]
def cluster_role_candidates(clusters, user_groups):
    """Calculate entitlement prevalence within each candidate user cluster."""
    candidates = []
    for cluster_id, users in enumerate(clusters, 1):
        counts = defaultdict(int)
        for user in users:
            for gid in user_groups[user]:
                counts[gid] += 1
        core = [
            (gid, count / len(users), count)
            for gid, count in counts.items()
            if count / len(users) >= ROLE_GROUP_PREVALENCE
        ]
        core.sort(key=lambda x: (-x[1], x[0]))
        if core:
            candidates.append((cluster_id, users, core))
    return candidates
def mine_frequent_patterns(user_groups):
    """Run sparse FP-Growth and derive high-confidence, high-lift rules."""
    transactions = [sorted(groups) for groups in user_groups.values() if groups]
    if len(transactions) < MIN_ROLE_USERS:
        return pd.DataFrame(), pd.DataFrame()
    # Near-universal groups carry no discriminative signal and are the primary
    # cause of combinatorial itemset/rule explosion. Drop them before mining.
    n_tx = len(transactions)
    freq = defaultdict(int)
    for t in transactions:
        for g in t:
            freq[g] += 1
    ubiquitous = {g for g, c in freq.items() if c / n_tx > DROP_UBIQUITOUS_ABOVE}
    if ubiquitous:
        transactions = [[g for g in t if g not in ubiquitous] for t in transactions]
        transactions = [t for t in transactions if t]
        print(
            f"  Dropped {len(ubiquitous)} ubiquitous group(s) "
            f"(>{DROP_UBIQUITOUS_ABOVE:.0%} of users) before FP-Growth"
        )
    if len(transactions) < MIN_ROLE_USERS:
        return pd.DataFrame(), pd.DataFrame()
    encoder = TransactionEncoder()
    matrix = encoder.fit(transactions).transform(transactions, sparse=True)
    frame = pd.DataFrame.sparse.from_spmatrix(matrix, columns=encoder.columns_)
    min_support = max(MIN_ITEMSET_SUPPORT, MIN_ROLE_USERS / len(transactions))
    all_itemsets = fpgrowth(
        frame,
        min_support=min_support,
        use_colnames=True,
        max_len=MAX_ITEMSET_LENGTH,
    )
    if all_itemsets.empty:
        return all_itemsets, pd.DataFrame()
    itemsets = all_itemsets[all_itemsets["itemsets"].map(len) >= 2].copy()
    itemsets["user_count"] = (
        itemsets["support"] * len(transactions)
    ).round().astype(int)
    # Some optional mlxtend metrics are undefined for ubiquitous groups; the
    # confidence and lift values used here remain valid.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", RuntimeWarning)
        rules = association_rules(
            all_itemsets,
            metric="confidence",
            min_threshold=MIN_RULE_CONFIDENCE,
        )
    if not rules.empty:
        rules = rules[rules["lift"] >= MIN_RULE_LIFT].copy()
        rules["user_count"] = (rules["support"] * len(transactions)).round().astype(int)
    # Hard-cap the outputs so a single sheet can never approach Excel's row
    # limit even in pathological tenants.
    itemsets = itemsets.sort_values(["support"], ascending=False).head(MAX_OUTPUT_ROWS)
    if not rules.empty:
        rules = rules.sort_values(["lift", "confidence"], ascending=False).head(MAX_OUTPUT_ROWS)
    return itemsets, rules

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
LOCK_FILL = PatternFill("solid", fgColor="E2EFDA")
PRIV_FILL = PatternFill("solid", fgColor="FFF2CC")   # privileged = amber
EXCEL_MAX_ROWS = 1_048_576
def pct(ws, *cols, fmt="0.0%"):
    """Apply a number format to data rows without ever exceeding Excel's limit."""
    last = min(ws.max_row, EXCEL_MAX_ROWS)
    for r in range(2, last + 1):
        for c in cols:
            ws.cell(r, c).number_format = fmt
def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18
def autosize(ws, maxw=60):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, maxw)
def owners_str(g):
    return "; ".join(g["owners"]) or "(no owner)"
def group_label(gid, data, duplicate_names=None):
    name = data[gid]["name"] or "(unnamed group)"
    if duplicate_names and name.casefold() in duplicate_names:
        return f"{name} [{gid[:8]}]"
    return name
def sanitize_excel(wb):
    """Prevent directory-controlled strings from becoming Excel formulas."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
                    cell.value = "'" + cell.value
def editable(g):
    if g["synced"]:
        return "No - AD synced"
    if g["dynamic"]:
        return "No - dynamic rule"
    return "Yes"
def build_workbook(data, identical, similar_groups, subsets, roles_available,
                   roles_note, similar_users, user_clusters, role_candidates,
                   frequent_itemsets, association_rule_rows, user_groups,
                   user_names):
    wb = Workbook()
    # ---- Groups ----
    ws = wb.active
    ws.title = "Groups"
    cols = ["Group", "Type", "Owners", "Members", "Editable?", "Teams?",
            "PrivilegedRole", "Synced", "Dynamic", "RoleAssignable", "Visibility",
            "LifecycleAgeDays", "Created", "Renewed", "Flags", "GroupId"]
    ws.append(cols)
    for gid, g in sorted(data.items(), key=lambda kv: kv[1]["name"].lower()):
        flags = cleanup_reasons(g)
        priv = role_summary(gid, data) if roles_available else ""
        row = [g["name"], g["type"], owners_str(g), len(g["members"]), editable(g),
               "Yes" if g["teams_connected"] else "",
               priv,
               "Yes" if g["synced"] else "", "Yes" if g["dynamic"] else "",
               "Yes" if g["role_assignable"] else "", g["visibility"],
               g["lifecycle_age_days"],
               g["created"].date().isoformat() if g["created"] else "",
               g["renewed"].date().isoformat() if g["renewed"] else "",
               ", ".join(flags), gid]
        ws.append(row)
        r = ws.max_row
        if priv:
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = PRIV_FILL
        elif flags:
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = WARN_FILL
        if g["synced"] or g["dynamic"] or g["teams_connected"]:
            ws.cell(r, 5).fill = LOCK_FILL
    style_header(ws, len(cols))
    autosize(ws)
    # ---- Members ----
    ws = wb.create_sheet("Members")
    ws.append(["Group", "Type", "Owners", "MemberName", "MemberType", "MemberId", "GroupId"])
    for gid, g in data.items():
        if not g["members"]:
            ws.append([g["name"], g["type"], owners_str(g), "(no members)", "", "", gid])
        for mid in g["members"]:
            mo = g["member_objs"].get(mid, {"name": mid, "type": ""})
            ws.append([g["name"], g["type"], owners_str(g), mo["name"], mo["type"], mid, gid])
    style_header(ws, 7)
    autosize(ws)
    # ---- Identical ----
    ws = wb.create_sheet("Identical")
    ws.append(["ClusterId", "Group", "Type", "Owners", "Editable?", "MemberCount", "GroupId"])
    for i, gids in enumerate(identical, 1):
        for gid in gids:
            g = data[gid]
            ws.append([i, g["name"], g["type"], owners_str(g), editable(g),
                       len(g["members"]), gid])
    style_header(ws, 7)
    autosize(ws)
    # ---- Similar ----
    ws = wb.create_sheet("Similar Groups")
    ws.append([
        "Similarity", "GroupA", "A_Members", "A_Owners", "A_Editable",
        "GroupB", "B_Members", "B_Owners", "B_Editable", "SharedMembers",
        "ratio"
    ])
    for id_a, id_b, s, shared in similar_groups:
        a, b = data[id_a], data[id_b]
        a_members = len(a["members"])
        b_members = len(b["members"])
        largest = max(a_members, b_members)
        ratio = min(a_members, b_members) / largest if largest else 0
        ws.append([
            round(s, 4),
            a["name"],
            a_members,
            owners_str(a),
            editable(a),
            b["name"],
            b_members,
            owners_str(b),
            editable(b),
            shared,
            ratio
        ])
    pct(ws, 1, 11)   # Similarity + Ratio
    style_header(ws, 11)
    autosize(ws)
    # ---- Subsets ----
    ws = wb.create_sheet("Subsets")
    ws.append(["SubsetGroup", "Subset_Members", "Subset_Owners",
               "SupersetGroup", "Superset_Members", "Superset_Owners", "Shared"])
    for id_sub, id_sup, shared in sorted(subsets, key=lambda x: x[2], reverse=True):
        sub, sup = data[id_sub], data[id_sup]
        ws.append([sub["name"], len(sub["members"]), owners_str(sub),
                   sup["name"], len(sup["members"]), owners_str(sup), shared])
    style_header(ws, 7)
    autosize(ws)
    # ---- Group Nesting (memberOf + nested children) ----
    ws = wb.create_sheet("Group Nesting")
    ws.append(["Group", "Type", "MemberOf (parent groups)", "ParentCount",
               "Nested child groups", "ChildCount", "GroupId"])
    for gid, g in sorted(data.items(), key=lambda kv: kv[1]["name"].lower()):
        parents = [data[p]["name"] for p in g["parents"] if p in data]
        children = [data[c]["name"] for c in g["child_groups"] if c in data]
        if not parents and not children:
            continue
        ws.append([g["name"], g["type"], "; ".join(parents), len(parents),
                   "; ".join(children), len(children), gid])
    style_header(ws, 7)
    autosize(ws)
    # ---- User Memberships (per-user -> groups) ----
    ws = wb.create_sheet("User Memberships")
    ws.append(["MemberName", "MemberType", "GroupCount", "Groups", "MemberId"])
    user_map = defaultdict(lambda: {"type": "", "groups": []})
    for gid, g in data.items():
        for mid, mo in g["member_objs"].items():
            user_map[mid]["type"] = mo["type"]
            user_map[mid]["groups"].append(g["name"])
    for mid, info in sorted(user_map.items(),
                            key=lambda kv: len(kv[1]["groups"]), reverse=True):
        # resolve a display name from any group that saw this member
        name = next((g["member_objs"][mid]["name"]
                     for g in data.values() if mid in g["member_objs"]), mid)
        ws.append([name, info["type"], len(info["groups"]),
                   "; ".join(sorted(info["groups"])), mid])
    style_header(ws, 5)
    autosize(ws)
    # Duplicate display names are disambiguated in all RBAC outputs.
    name_counts = defaultdict(int)
    for g in data.values():
        name_counts[g["name"].casefold()] += 1
    duplicate_names = {name for name, count in name_counts.items() if count > 1}
    # ---- Similar users (Jaccard pairs) ----
    ws = wb.create_sheet("Similar Users")
    ws.append([
        "Similarity",
        "UserA",
        "A_Groups",
        "UserB",
        "B_Groups",
        "SharedGroups",
        "OnlyA",
        "OnlyB"
    ])
    for u1, u2, sim, shared, only_a, only_b in similar_users:
        ws.append([
            sim,
            user_names.get(u1, u1),
            len(user_groups[u1]),
            user_names.get(u2, u2),
            len(user_groups[u2]),
            len(shared),
            "; ".join(sorted(group_label(g, data, duplicate_names) for g in only_a)),
            "; ".join(sorted(group_label(g, data, duplicate_names) for g in only_b)),
        ])
    pct(ws, 1)
    style_header(ws, 8)
    autosize(ws)
    # ---- Jaccard user clusters ----
    ws = wb.create_sheet("User Clusters")
    ws.append(["ClusterId", "Users", "User", "GroupCount", "Groups", "UserId"])
    for cluster_id, cluster in enumerate(user_clusters, 1):
        for uid in cluster:
            labels = sorted(group_label(g, data, duplicate_names) for g in user_groups[uid])
            ws.append([cluster_id, len(cluster), user_names.get(uid, uid),
                       len(labels), "; ".join(labels), uid])
    style_header(ws, 6)
    autosize(ws)
    # ---- Role candidates derived from cluster prevalence ----
    ws = wb.create_sheet("Role Candidates")
    ws.append(["ClusterId", "ClusterUsers", "CandidateGroup", "Prevalence",
               "UsersWithGroup", "GroupId"])
    for cluster_id, users, core in role_candidates:
        for gid, prevalence, count in core:
            ws.append([cluster_id, len(users), group_label(gid, data, duplicate_names),
                       prevalence, count, gid])
    pct(ws, 4)
    style_header(ws, 6)
    autosize(ws)
    # ---- FP-Growth bundles ----
    ws = wb.create_sheet("Frequent Bundles")
    ws.append(["Groups", "GroupCount", "Users", "Support"])
    if frequent_itemsets is not None and not frequent_itemsets.empty:
        for _, row in frequent_itemsets.iterrows():
            gids = row["itemsets"]
            labels = sorted(group_label(g, data, duplicate_names) for g in gids)
            ws.append(["; ".join(labels), len(gids), int(row["user_count"]),
                       float(row["support"])])
    pct(ws, 4)
    style_header(ws, 4)
    autosize(ws)
    # ---- Association rules ----
    ws = wb.create_sheet("Association Rules")
    ws.append(["ExistingGroups", "PredictedGroups", "Users", "Support",
               "Confidence", "Lift", "MissingPredictionUsers"])
    if association_rule_rows is not None and not association_rule_rows.empty:
        for _, row in association_rule_rows.sort_values(
                ["lift", "confidence"], ascending=False).iterrows():
            antecedents = set(row["antecedents"])
            consequents = set(row["consequents"])
            matching = [uid for uid, groups in user_groups.items()
                        if antecedents <= groups]
            missing = [user_names.get(uid, uid) for uid in matching
                       if not consequents <= user_groups[uid]]
            ws.append([
                "; ".join(sorted(group_label(g, data, duplicate_names)
                                 for g in antecedents)),
                "; ".join(sorted(group_label(g, data, duplicate_names)
                                 for g in consequents)),
                int(row["user_count"]), float(row["support"]),
                float(row["confidence"]), float(row["lift"]),
                "; ".join(sorted(missing)),
            ])
    pct(ws, 4, 5)
    pct(ws, 6, fmt="0.00")
    style_header(ws, 7)
    autosize(ws)
    # ---- Privileged Groups ----
    ws = wb.create_sheet("Privileged Groups")
    if not roles_available:
        ws.append(["Role data unavailable"])
        ws.append([roles_note])
    else:
        ws.append(["Group", "Role", "Assignment", "Source", "Members", "Owners", "GroupId"])
        any_priv = False
        for gid, g in sorted(data.items(), key=lambda kv: kv[1]["name"].lower()):
            for rn, kind, src in effective_roles(gid, data):
                any_priv = True
                ws.append([g["name"], rn, kind, src, len(g["members"]),
                           owners_str(g), gid])
                for c in range(1, 8):
                    ws.cell(ws.max_row, c).fill = PRIV_FILL
        if not any_priv:
            ws.append(["(no groups are assigned or eligible for directory roles)"])
        style_header(ws, 7)
    autosize(ws)
    # ---- Cleanup Targets (exclude Teams-connected AND privileged) ----
    ws = wb.create_sheet("Cleanup Targets")
    ws.append(["Priority", "Group", "Type", "Owners", "Members",
               "Editable?", "LifecycleAgeDays", "Reasons", "GroupId"])
    targets = []
    excluded = 0
    for gid, g in data.items():
        reasons = cleanup_reasons(g)
        if not reasons:
            continue
        if g["teams_connected"] or (roles_available and is_privileged(gid, data)):
            excluded += 1
            continue
        score = (("No owner" in reasons) + ("Empty" in reasons) * 2
                 + any("Lifecycle review" in x for x in reasons))
        targets.append((score, gid, reasons))
    for score, gid, reasons in sorted(targets, key=lambda x: x[0], reverse=True):
        g = data[gid]
        ws.append([score, g["name"], g["type"], owners_str(g), len(g["members"]),
                   editable(g), g["lifecycle_age_days"], ", ".join(reasons), gid])
        for c in range(1, 10):
            ws.cell(ws.max_row, c).fill = WARN_FILL
    style_header(ws, 9)
    autosize(ws)
    # ---- Summary (first) ----
    ws = wb.create_sheet("Summary", 0)
    ownerless = sum(1 for g in data.values() if not g["owners"])
    empty = sum(1 for g in data.values() if not g["members"])
    lifecycle_review = sum(1 for g in data.values() if needs_lifecycle_review(g))
    synced = sum(1 for g in data.values() if g["synced"])
    dynamic = sum(1 for g in data.values() if g["dynamic"])
    teams = sum(1 for g in data.values() if g["teams_connected"])
    priv = sum(1 for gid in data if roles_available and is_privileged(gid, data))
    rows = [
        ["Entra ID Group Cleanup Report", ""],
        ["Generated (UTC)", NOW.strftime("%Y-%m-%d %H:%M")],
        ["Similarity metric", SIMILARITY_METRIC],
        ["Similarity threshold", f"{SIMILARITY_THRESHOLD:.0%}"],
        ["Lifecycle review threshold", LIFECYCLE_REVIEW_DAYS],
        ["Lifecycle warning", "Created/renewed age is not evidence of inactivity."],
        ["Membership scope", "Transitive" if TRANSITIVE else "Direct members only"],
        ["RBAC group scope", "Security-enabled groups only" if RBAC_SECURITY_GROUPS_ONLY else "All groups"],
        ["User Jaccard threshold", f"{USER_JACCARD_THRESHOLD:.0%}"],
        ["Minimum shared groups", MIN_SHARED_GROUPS],
        ["Minimum role users", MIN_ROLE_USERS],
        ["Role-group prevalence", f"{ROLE_GROUP_PREVALENCE:.0%}"],
        ["FP-Growth confidence / lift", f"{MIN_RULE_CONFIDENCE:.0%} / {MIN_RULE_LIFT:.2f}"],
        ["Drop-ubiquitous threshold", f"{DROP_UBIQUITOUS_ABOVE:.0%}"],
        ["Max output rows (bundles/rules)", MAX_OUTPUT_ROWS],
        ["Role data", "Loaded" if roles_available else roles_note],
        ["", ""],
        ["Total groups", len(data)],
        ["Identical-member clusters", len(identical)],
        [f"Similar group pairs (>= {SIMILARITY_THRESHOLD:.0%})", len(similar_groups)],
        [f"Similar user pairs (>= {USER_JACCARD_THRESHOLD:.0%})", len(similar_users)],
        ["Candidate user clusters", len(user_clusters)],
        ["Candidate role clusters", len(role_candidates)],
        ["Frequent group bundles", len(frequent_itemsets) if frequent_itemsets is not None else 0],
        ["Association rules", len(association_rule_rows) if association_rule_rows is not None else 0],
        ["Subset/containment pairs", len(subsets)],
        ["Teams-connected groups", teams],
        ["Privileged groups (Active/Eligible)", priv if roles_available else "n/a"],
        ["Ownerless groups", ownerless],
        ["Empty groups", empty],
        ["Groups needing lifecycle review", lifecycle_review],
        ["AD-synced (not editable in Entra)", synced],
        ["Dynamic groups (rule-managed)", dynamic],
        ["", ""],
        ["How to read this workbook", ""],
        ["Cleanup Targets", "Start here. Teams-connected and privileged groups are EXCLUDED."],
        ["Privileged Groups", "Groups holding Entra roles (Active or Eligible-PIM), incl. inherited."],
        ["Group Nesting", "Parent groups (memberOf) + nested child groups per group."],
        ["User Memberships", "Per-user view: which groups each member belongs to."],
        ["Similar Users", "Jaccard pairs meeting similarity and minimum-shared-group thresholds."],
        ["User Clusters", "Complete-link-style Jaccard clusters; every member matches every other."],
        ["Role Candidates", "Groups present in the configured percentage of each user cluster."],
        ["Frequent Bundles", "FP-Growth group bundles meeting minimum user/support thresholds."],
        ["Association Rules", "High-confidence/high-lift relationships plus missing predictions."],
        ["Identical / Similar / Subsets", "Merge/dedup candidates; 'Editable?' shows what you can touch."],
        ["Amber rows / cells", "Privileged - verify before ANY change."],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 62
    # Final guard: no sheet may exceed Excel's hard row limit.
    for s in wb.worksheets:
        if s.max_row > EXCEL_MAX_ROWS:
            print(f"OVERFLOW: '{s.title}' has {s.max_row} rows (Excel max {EXCEL_MAX_ROWS})")
    sanitize_excel(wb)
    wb.save(OUTFILE)
# ---------------------------------------------------------------------------
def main():
    headers = {"Authorization": f"Bearer {get_token()}"}
    print("Collecting groups, members, owners...")
    data = collect_groups(headers)
    build_nesting(data)
    print("Collecting privileged-role assignments...")
    roles_available, roles_note = collect_roles(headers, data)
    if not roles_available:
        print(f"  WARNING: {roles_note}")
    identical = find_identical(data)
    similar_groups, subsets = find_similar_and_subsets(
        data, SIMILARITY_THRESHOLD, SIMILARITY_METRIC
    )
    print("Discovering RBAC patterns...")
    user_groups, user_names = build_user_group_sets(data)
    similar_users = find_similar_users(user_groups)
    user_clusters = cluster_similar_users(user_groups, similar_users)
    role_candidates = cluster_role_candidates(user_clusters, user_groups)
    frequent_itemsets, association_rule_rows = mine_frequent_patterns(user_groups)
    build_workbook(
        data, identical, similar_groups, subsets, roles_available, roles_note,
        similar_users, user_clusters, role_candidates, frequent_itemsets,
        association_rule_rows, user_groups, user_names,
    )
    print(f"\nTotal groups: {len(data)}")
    print(
        f"Identical: {len(identical)} | Similar groups: {len(similar_groups)} "
        f"| Subsets: {len(subsets)}"
    )
    print(
        f"Similar users: {len(similar_users)} | User clusters: {len(user_clusters)} "
        f"| Role candidates: {len(role_candidates)}"
    )
    print(f"Teams-connected: {sum(1 for g in data.values() if g['teams_connected'])}")
    if roles_available:
        print(f"Privileged: {sum(1 for gid in data if is_privileged(gid, data))}")
    print(f"\nWorkbook written: {OUTFILE}")
if __name__ == "__main__":
    main()
